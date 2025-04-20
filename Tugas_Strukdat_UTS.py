# Import Library
import re
import os
import platform
from openpyxl import load_workbook  

# Load file Excel
file_path = "Struktur_Data_Dataset_Kelas_A_B_C.xlsx"
wb = load_workbook(file_path)
sheet = wb.active  # ambil sheet pertama

# Ambil data dari kolom F-K (index 5 sampai 10)
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
    if row[5] and row[6] and row[7]:  # pastikan minimal kolom penting tidak kosong
        data.append({
            "No": str(int(row[0])),
            "NIM": str(int(row[1])),
            "Nama Mahasiswa": str(row[2]),
            "Sumber Database":str(row[3]),
            "Fokus Kata Kunci":str(row[4]),
            "Judul Paper": str(row[5]),
            "Tahun Terbit": int(row[6]) if isinstance(row[6], float) else row[6],
            "Nama Penulis": str(row[7]),
            "Abstrak": str(row[8]) if len(row) > 8 and row[8] else "Tidak tersedia",
            "Kesimpulan": str(row[9]) if len(row) > 9 and row[9] else "Tidak tersedia",
            "Link Paper": str(row[10]) if len(row) > 10 and row[10] else "Tidak tersedia"
        })

# Fungsi untuk membersihkan layar
def clear_screen():
    os.system('cls' if platform.system() == 'Windows' else 'clear')

# Fungsi Linear Search
def linear_search(data, key, attribute):
    results = []
    pattern = r'\b' + re.escape(key.lower()) + r'\b'
    for item in data:
        if re.search(pattern, str(item[attribute]).lower()):
            results.append(item)
    return results

# Fungsi Binary Search
def binary_search(data, key, attribute):
    sorted_data = sorted(data, key=lambda x: str(x[attribute]).lower())
    low = 0
    high = len(sorted_data) - 1
    results = []
    pattern = r'\b' + re.escape(key.lower()) + r'\b'

    while low <= high:
        mid = (low + high) // 2
        mid_value = str(sorted_data[mid][attribute]).lower()

        if re.search(pattern, mid_value):
            results.append(sorted_data[mid])
            left = mid - 1
            while left >= 0 and re.search(pattern, str(sorted_data[left][attribute]).lower()):
                results.append(sorted_data[left])
                left -= 1
            right = mid + 1
            while right < len(sorted_data) and re.search(pattern, str(sorted_data[right][attribute]).lower()):
                results.append(sorted_data[right])
                right += 1
            return results
        elif key.lower() < mid_value:
            high = mid - 1
        else:
            low = mid + 1
    return results

# Fungsi menu program
def main():
    while True:
        clear_screen()  # Bersihkan layar setiap mulai ulang
        print("\n=== MENU PENCARIAN PAPER ===")
        print("1. Linear Search")
        print("2. Binary Search")
        choice = input("Masukkan pilihan (1/2): ")

        if choice not in ["1", "2"]:
            print("Pilihan tidak valid.")
            input("Tekan Enter untuk kembali...")
            continue

        print("\nPilih atribut pencarian:")
        print("1. Judul Paper")
        print("2. Tahun Terbit")
        print("3. Nama Penulis")
        while True:
            attribute_choice = input("Masukkan pilihan (1/2/3): ")
            if attribute_choice in ["1", "2", "3"]:
                break
            else:
                print("Pilihan atribut tidak valid. Silakan masukkan 1, 2, atau 3.")


        key = input("Masukkan kata kunci pencarian: ")

        attributes = {"1": "Judul Paper", "2": "Tahun Terbit", "3": "Nama Penulis"}
        attribute = attributes.get(attribute_choice, "Judul Paper")

        if choice == "1":
            results = linear_search(data, key, attribute)
        else:
            results = binary_search(data, key, attribute)

        clear_screen()  # Bersihkan layar sebelum tampilkan hasil
        if results:
            print("\nHasil Pencarian:")
            for result in results:
                print(f"No                  : {result['No']}")
                print(f"NIM                 : {result['NIM']}")
                print(f"Nama Mahasiswa      : {result['Nama Mahasiswa']}")
                print(f"Sumber Database     : {result['Sumber Database']}")
                print(f"Fokus Kata Kunci    : {result['Fokus Kata Kunci']}")
                print(f"Judul Paper         : {result['Judul Paper']}")
                print(f"Tahun Terbit        : {result['Tahun Terbit']}")
                print(f"Nama Penulis        : {result['Nama Penulis']}")
                print("\nAbstrak:")
                print(result["Abstrak"])
                print("\nKesimpulan:")
                print(result["Kesimpulan"])
                print(f"\nLink      : {result['Link Paper']}")
                print("\n" + "-" * 40 + "\n")
        else:
            print("Tidak ditemukan hasil yang sesuai.")

        keluar = input("\nApakah ingin keluar? (y/n): ").lower()
        if keluar == "y":
            print("Terima kasih! Program selesai.")
            break

if __name__ == "__main__":
    main()
